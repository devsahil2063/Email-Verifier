import pandas as pd
import re
import io
from typing import List, Optional
import openpyxl
from openpyxl.styles import PatternFill, Font


class ExcelProcessor:
    """
    Excel file processor for handling email verification workflows
    """
    
    def __init__(self):
        """Initialize Excel processor"""
        self.email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    
    def read_excel(self, file_path_or_buffer) -> pd.DataFrame:
        """
        Read Excel file and return DataFrame
        
        Args:
            file_path_or_buffer: File path or buffer object
            
        Returns:
            pd.DataFrame: Loaded data
        """
        try:
            # Try to read as .xlsx first, then .xls
            try:
                df = pd.read_excel(file_path_or_buffer, engine='openpyxl')
            except:
                df = pd.read_excel(file_path_or_buffer, engine='xlrd')
            
            # Clean column names
            df.columns = df.columns.astype(str)
            
            # Remove completely empty rows
            df = df.dropna(how='all')
            
            return df
            
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    def detect_email_columns(self, df: pd.DataFrame) -> List[str]:
        """
        Auto-detect columns that likely contain email addresses
        
        Args:
            df: Input DataFrame
            
        Returns:
            List[str]: Column names that likely contain emails
        """
        email_columns = []
        
        for column in df.columns:
            # Check column name for email-related keywords
            column_lower = str(column).lower()
            if any(keyword in column_lower for keyword in ['email', 'mail', 'e-mail', '@']):
                email_columns.append(column)
                continue
            
            # Check actual data in column for email patterns
            sample_size = min(10, len(df))
            sample_data = df[column].dropna().head(sample_size)
            
            email_count = 0
            for value in sample_data:
                if self.is_valid_email_format(str(value)):
                    email_count += 1
            
            # If more than 30% of samples are valid emails, consider it an email column
            if sample_size > 0 and (email_count / sample_size) > 0.3:
                email_columns.append(column)
        
        return email_columns
    
    def is_valid_email_format(self, email: str) -> bool:
        """
        Check if string matches email format
        
        Args:
            email: String to check
            
        Returns:
            bool: True if valid email format
        """
        try:
            return bool(re.match(self.email_pattern, email.strip()))
        except:
            return False
    
    def extract_valid_emails(self, df: pd.DataFrame, email_column: str) -> List[str]:
        """
        Extract valid email addresses from specified column
        
        Args:
            df: Input DataFrame
            email_column: Column name containing emails
            
        Returns:
            List[str]: Valid email addresses
        """
        if email_column not in df.columns:
            return []
        
        valid_emails = []
        for email in df[email_column].dropna():
            email_str = str(email).strip()
            if self.is_valid_email_format(email_str):
                valid_emails.append(email_str)
        
        return valid_emails
    
    def add_verification_columns(self, df: pd.DataFrame, results: dict) -> pd.DataFrame:
        """
        Add verification results to DataFrame
        
        Args:
            df: Original DataFrame
            results: Dictionary with email verification results
            
        Returns:
            pd.DataFrame: DataFrame with verification columns added
        """
        result_df = df.copy()
        
        # Add verification columns if they don't exist
        if 'Email_Verification_Status' not in result_df.columns:
            result_df['Email_Verification_Status'] = 'Not Checked'
        
        if 'Verification_Details' not in result_df.columns:
            result_df['Verification_Details'] = ''
        
        # Update results
        for email, result in results.items():
            # Find rows with this email
            mask = result_df.apply(lambda row: str(email) in row.values, axis=1)
            result_df.loc[mask, 'Email_Verification_Status'] = 'Valid' if result['is_valid'] else 'Invalid'
            result_df.loc[mask, 'Verification_Details'] = result['details']
        
        return result_df
    
    def dataframe_to_excel(self, df: pd.DataFrame, include_formatting: bool = True) -> bytes:
        """
        Convert DataFrame to Excel file with optional formatting
        
        Args:
            df: DataFrame to convert
            include_formatting: Whether to include color formatting
            
        Returns:
            bytes: Excel file as bytes
        """
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Email_Verification_Results')
            
            if include_formatting and 'Email_Verification_Status' in df.columns:
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Email_Verification_Results']
                
                # Define colors for different statuses
                valid_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
                invalid_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light red
                error_fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')  # Light orange
                
                # Find the status column
                status_col_index = None
                for idx, col in enumerate(df.columns, 1):
                    if col == 'Email_Verification_Status':
                        status_col_index = idx
                        break
                
                if status_col_index:
                    # Apply formatting to status column
                    for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                        cell = worksheet.cell(row=row_idx, column=status_col_index)
                        if cell.value == 'Valid':
                            cell.fill = valid_fill
                        elif cell.value == 'Invalid':
                            cell.fill = invalid_fill
                        elif cell.value == 'Error':
                            cell.fill = error_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def get_file_stats(self, df: pd.DataFrame, email_column: Optional[str] = None) -> dict:
        """
        Get statistics about the DataFrame and email data
        
        Args:
            df: Input DataFrame
            email_column: Email column name
            
        Returns:
            dict: File statistics
        """
        stats = {
            'total_rows': len(df),
            'total_columns': len(df.columns),
            'columns': df.columns.tolist(),
            'empty_rows': df.isnull().all(axis=1).sum()
        }
        
        if email_column and email_column in df.columns:
            email_series = df[email_column].dropna()
            valid_emails = [email for email in email_series if self.is_valid_email_format(str(email))]
            
            stats.update({
                'total_emails': len(email_series),
                'valid_emails': len(valid_emails),
                'invalid_emails': len(email_series) - len(valid_emails),
                'empty_emails': df[email_column].isnull().sum()
            })
        
        return stats
    
    def clean_email_data(self, df: pd.DataFrame, email_column: str) -> pd.DataFrame:
        """
        Clean email data in the specified column
        
        Args:
            df: Input DataFrame
            email_column: Column containing emails
            
        Returns:
            pd.DataFrame: DataFrame with cleaned email data
        """
        if email_column not in df.columns:
            return df
        
        cleaned_df = df.copy()
        
        # Clean email strings
        cleaned_df[email_column] = cleaned_df[email_column].apply(
            lambda x: str(x).strip().lower() if pd.notna(x) else x
        )
        
        # Remove obviously invalid entries
        cleaned_df[email_column] = cleaned_df[email_column].apply(
            lambda x: x if (pd.notna(x) and '@' in str(x) and '.' in str(x)) else None
        )
        
        return cleaned_df
    
    def create_summary_report(self, df: pd.DataFrame) -> dict:
        """
        Create a summary report of verification results
        
        Args:
            df: DataFrame with verification results
            
        Returns:
            dict: Summary report
        """
        if 'Email_Verification_Status' not in df.columns:
            return {}
        
        status_counts = df['Email_Verification_Status'].value_counts().to_dict()
        total_checked = len(df[df['Email_Verification_Status'] != 'Not Checked'])
        
        return {
            'total_rows': len(df),
            'total_checked': total_checked,
            'valid_emails': status_counts.get('Valid', 0),
            'invalid_emails': status_counts.get('Invalid', 0),
            'errors': status_counts.get('Error', 0),
            'not_checked': status_counts.get('Not Checked', 0),
            'invalid_format': status_counts.get('Invalid Format', 0),
            'success_rate': (status_counts.get('Valid', 0) / total_checked * 100) if total_checked > 0 else 0
        }
